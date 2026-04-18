from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.models import CurriculumPlan, PlanElement
from backend.modules.plan_builder.calculator import aggregate_by_block


BLOCK_TITLES = {
    "1": "Блок 1. Дисциплины",
    "2": "Блок 2. Практики",
    "3": "Блок 3. ГИА",
    "fac": "Факультативы",
}

PART_TITLES = {
    "mandatory": "Обязательная часть",
    "variative": "Вариативная часть",
}


def _format_semesters(semesters: list[int]) -> str:
    values = sorted(int(item) for item in semesters or [])
    return ", ".join(str(item) for item in values)


def build_plan_workbook(plan: CurriculumPlan, elements: list[PlanElement]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Учебный план"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)

    sheet["A1"] = "Учебный план"
    sheet["A1"].font = title_font
    sheet["A2"] = f"План: {plan.name}"
    sheet["A3"] = f"Статус: {plan.status}"

    header_row = 5
    headers = [
        "Наименование",
        "Блок",
        "Часть",
        "Семестры",
        "З.е.",
        "Часы",
        "Компетенции",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = header_font

    current_row = header_row + 1
    current_group: tuple[str, str] | None = None

    for element in elements:
        group = (str(element.block), str(element.part))
        if group != current_group:
            block_title = BLOCK_TITLES.get(group[0], f"Блок {group[0]}")
            part_title = PART_TITLES.get(group[1], group[1])
            sheet.cell(row=current_row, column=1, value=block_title)
            sheet.cell(row=current_row, column=2, value=part_title)
            sheet.cell(row=current_row, column=1).font = header_font
            sheet.cell(row=current_row, column=2).font = header_font
            current_row += 1
            current_group = group

        competency_text = ", ".join(str(item) for item in element.competency_ids)
        row_values = [
            element.name,
            str(element.block),
            str(element.part),
            _format_semesters(element.semesters),
            element.credits,
            element.hours,
            competency_text,
        ]
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=current_row, column=column_index, value=value)
        current_row += 1

    current_row += 1
    sheet.cell(row=current_row, column=1, value="Итоги по блокам").font = header_font
    current_row += 1
    for block, total in aggregate_by_block(elements).items():
        sheet.cell(row=current_row, column=1, value=BLOCK_TITLES.get(block, f"Блок {block}"))
        sheet.cell(row=current_row, column=2, value=total)
        current_row += 1

    column_widths = {
        "A": 45,
        "B": 18,
        "C": 22,
        "D": 16,
        "E": 10,
        "F": 10,
        "G": 24,
    }
    for column_name, width in column_widths.items():
        sheet.column_dimensions[column_name].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
